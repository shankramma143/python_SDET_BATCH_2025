import openpyxl
file="C:\\Users\\Satish\\Downloads\\excel sample.xlsx"
workbook=openpyxl.load_workbook(file)


#To get the particular sheet
sheet=workbook['Sheet1']


# Get total number of rows count
rows=sheet.max_row
print(rows)


# Get total number of column count
column=sheet.max_column
print(column)


#Read all the rows and Columns in the excel file
for r in range(1,rows+1):
   for c in range(1,column+1):
       print(sheet.cell(r,c).value ,end='  ')
   print()
