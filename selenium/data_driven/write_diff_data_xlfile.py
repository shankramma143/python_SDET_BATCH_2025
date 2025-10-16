import openpyxl
file="C:\\Users\\Satish\\Downloads\\datadriven write.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active


# Get total number of rows count
rows=sheet.max_row
print(rows)


# Get total number of column count
column=sheet.max_column
print(column)

sheet.cell(2,1).value="1"
sheet.cell(2,2).value="kavita"

sheet.cell(3,1).value="2"
sheet.cell(3,2).value="rajitha"

sheet.cell(4,1).value="3"
sheet.cell(4,2).value="shankru"

workbook.save(file)

# Get total number of rows count
rows=sheet.max_row
print(rows)


# Get total number of column count
column=sheet.max_column
print(column)

